import requests
import json
import sys
import pandas as pd
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

args = sys.argv
args_len = len(sys.argv)

k_arguments = ['rnr', 'gruppe']

current_time = datetime.now().date()

"""Parse command line arguments"""
def parse_args(args):

    global k_arguments
    global color_flag 

    color_flag = True   # Default value
    
    for i in range (1, args_len):
        if (args[i] == "-k" or args[i] == "--keys"):
           k_arguments = k_arguments + args[i+1].split(',')
           
        if (args[i] == "-c" or args[i] == "--colored"):
            color_flag = eval(args[i+1].lower().capitalize()) is True


"""First login to the API"""
def login():
    global acc_tkn

    payload = {
        "username": "365",
        "password": "1"
    }

    url = "https://api.baubuddy.de/index.php/login"

    headers = {
        "Authorization": "Basic QVBJX0V4cGxvcmVyOjEyMzQ1NmlzQUxhbWVQYXNz",
        "Content-Type": "application/json"
    }

    response = requests.request("POST", url, json=payload, headers=headers)
    acc_tkn = response.json()['oauth']['access_token']


"""Getting the Veachles data from API"""
def request_resourse(acc_tkn):
    global rq

    url = 'https://api.baubuddy.de/dev/index.php/v1/vehicles/select/active'
    
    headers = {
        "Authorization": f"Bearer {acc_tkn}",
        "Content-Type": "application/json"
    }

    rq = requests.request('GET', url=url, headers=headers).json()


'''Filtering out any resources that do not have a value set for hu field'''
def filter(rq):
    global new_list
    new_list = []
    for char in rq:
        if char['hu'] == None:
            pass
        else:
            new_list.append(char)
    return new_list
    

'''1. Sorting every rows by field gruppe
   2. Showing in the excel file only those column that are written to the argument -k'''
def sort_excel():
    jsonStr = json.dumps(filter(rq), indent=2)
    xlsx_file = pd.read_json(jsonStr)
    xlsx_file.sort_values('gruppe').to_excel(f'vehicles_{current_time}_iso_formatted.xlsx', index=None, header=True)
    cols = pd.read_excel(f'vehicles_{current_time}_iso_formatted.xlsx')
    
    cols.drop(cols.columns.difference(k_arguments), axis=1, inplace=True)
    
    cols.to_excel(f'vehicles_{current_time}_iso_formatted.xlsx', index=None, header=True)
    
    if 'labelIds' in k_arguments:
        get_color_code()


def get_color_code():
    global names
    '''get the data from column rnr'''
    cols = pd.read_excel(f'vehicles_{current_time}_iso_formatted.xlsx')
    names = cols['rnr'].to_list()

    headers = {
        "Authorization": f"Bearer {acc_tkn}",
        "Content-Type": "application/json"
    }

    '''Get every colorCode into list'''
    color_list = []
    for i in range(len(names)):
        url = f'https://api.baubuddy.de/dev/index.php/v1/labels/{names[i]}'
        rq = requests.request('GET', url=url, headers=headers).json()
        if rq == []:
            color_list.append('')
        else:
            if 'colorCode' in rq[0] and rq[0]['colorCode'] != '':
                color_list.append(rq[0]['colorCode'])
            elif 'colorCode' not in rq[0] or rq[0]['colorCode'] == '':
                color_list.append('')

    '''append list of colorCodes into excel file'''
    cols['labelIds'] = color_list

    cols.to_excel(f'vehicles_{current_time}_iso_formatted.xlsx', index=None, header=True)


def color_rows():
    if color_flag == True:
        '''time deltas'''
        delta1 = timedelta(days=90)
        delta2 = timedelta(days=365)
        counter = 0
        color_fill = ''

        file = pd.read_excel(f'vehicles_{current_time}_iso_formatted.xlsx', sheet_name='Sheet1')

        with pd.ExcelWriter(f'vehicles_{current_time}_iso_formatted.xlsx', engine='openpyxl') as writer:
            file.to_excel(writer, sheet_name='Sheet1', index=False)
            sheet = writer.sheets['Sheet1']
            for row in sheet.iter_rows(min_row=2, max_row=(len(file) + 1), min_col=1):
                date_diff =  current_time - datetime.strptime(file['hu'][counter], '%Y-%m-%d').date()
                if date_diff < delta1:
                    color_fill = '007500'
                elif date_diff >= delta1 and date_diff < delta2:
                    color_fill = 'FFA500'
                elif date_diff >= delta2:
                    color_fill = 'b30000'
                counter = counter + 1
                for cell in row:
                    cell.fill = PatternFill('solid', start_color=(f'{color_fill}'))


def main():
    parse_args(args)
    login()
    request_resourse(acc_tkn)
    filter(rq)
    sort_excel()
    if 'hu' in k_arguments:
        color_rows()


if __name__ == '__main__':
    main()